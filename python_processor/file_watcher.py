"""
File Watcher Module
Monitors a folder for new PDF files and automatically processes them
"""

import time
import logging
import uuid
from pathlib import Path
from datetime import datetime

from watchdog.observers import Observer
from watchdog.events import FileSystemEventHandler, FileCreatedEvent

from config import config
from orchestrator import DocumentOrchestrator, ProcessRequest

logger = logging.getLogger(__name__)


class PDFHandler(FileSystemEventHandler):
    """Handler for PDF file events"""

    def __init__(self, doc_category: int = 1):
        super().__init__()
        self.doc_category = doc_category
        self.orchestrator = DocumentOrchestrator()
        self.processed_files = set()

    def on_created(self, event):
        """Handle file creation events"""
        if not isinstance(event, FileCreatedEvent):
            return

        if event.is_directory:
            return

        file_path = Path(event.src_path)

        # Only process PDF files
        if file_path.suffix.lower() != '.pdf':
            return

        # Skip already processed files
        if file_path.name.startswith('Processed_'):
            return

        if str(file_path) in self.processed_files:
            return

        logger.info(f"New PDF detected: {file_path.name}")

        # Wait for file to be fully written
        self._wait_for_file(file_path)

        # Process the file
        self._process_file(file_path)

    def _wait_for_file(self, file_path: Path, timeout: int = 30):
        """Wait for file to be completely written"""
        last_size = -1
        stable_count = 0

        start_time = time.time()

        while time.time() - start_time < timeout:
            try:
                current_size = file_path.stat().st_size
                if current_size == last_size and current_size > 0:
                    stable_count += 1
                    if stable_count >= 3:
                        return
                else:
                    stable_count = 0
                    last_size = current_size
            except Exception:
                pass

            time.sleep(1)

        logger.warning(f"Timeout waiting for file: {file_path.name}")

    def _process_file(self, file_path: Path):
        """Process a PDF file using configured providers"""
        self.processed_files.add(str(file_path))

        process_id = str(uuid.uuid4())[:8]

        logger.info(f"Processing file: {file_path.name} (ID: {process_id})")
        logger.info(f"  Using providers: {config.get_provider_info()}")

        try:
            import json
            from cost_tracker import CostTracker

            cost_tracker = CostTracker()
            cost_tracker.start_tracking()

            # Check which provider to use
            use_mistral_ocr = config.providers.ocr_provider == 'mistral'
            use_mistral_llm = config.providers.llm_provider == 'mistral'

            if use_mistral_ocr and use_mistral_llm:
                # Use Mistral for both OCR and extraction (most efficient)
                extraction_result, total_pages = self._process_with_mistral(file_path)
                provider_info = 'Mistral (OCR + LLM)'
            elif use_mistral_ocr:
                # Mistral OCR + OpenAI extraction
                extraction_result, total_pages = self._process_mistral_ocr_openai_llm(file_path)
                provider_info = 'Mistral OCR + OpenAI LLM'
            else:
                # Google Document AI + configured LLM
                extraction_result, total_pages = self._process_with_google_docai(file_path, use_mistral_llm)
                provider_info = f'Google DocAI + {"Mistral" if use_mistral_llm else "OpenAI"} LLM'

            if extraction_result and extraction_result.records:
                # Calculate costs based on provider
                cost_breakdown = cost_tracker.calculate_total_cost(
                    pages=total_pages,
                    input_tokens=extraction_result.input_tokens,
                    output_tokens=extraction_result.output_tokens,
                    provider=config.providers.llm_provider
                )

                logger.info(f"Processing complete: {file_path.name}")
                logger.info(f"  Provider: {provider_info}")
                logger.info(f"  Records: {len(extraction_result.records)}")
                logger.info(f"  Total Cost: ${cost_breakdown.total_cost:.4f}")
                logger.info(f"  Processing Time: {cost_breakdown.processing_time_seconds}s")

                # Save extraction result as JSON
                output_path = Path(config.folders.results_folder) / f"extracted_{file_path.stem}.json"
                with open(output_path, 'w', encoding='utf-8') as f:
                    json.dump({
                        'data': extraction_result.data,
                        'records': extraction_result.records,
                        'metadata': {
                            'process_id': process_id,
                            'filename': file_path.name,
                            'pages': total_pages,
                            'total_records': len(extraction_result.records),
                            'provider': provider_info,
                            'costs': {
                                'ocr': cost_breakdown.document_ai_cost,
                                'llm': cost_breakdown.openai_cost,
                                'total': cost_breakdown.total_cost
                            },
                            'processing_time_seconds': cost_breakdown.processing_time_seconds,
                            'processed_at': datetime.now().isoformat()
                        }
                    }, f, indent=2, ensure_ascii=False)
                logger.info(f"  JSON saved: {output_path.name}")

                # Save extraction result as CSV
                csv_path = Path(config.folders.results_folder) / f"extracted_{file_path.stem}.csv"
                self._save_records_to_csv(extraction_result.records, csv_path)
                logger.info(f"  CSV saved: {csv_path.name}")

                # Save extraction result as Excel
                try:
                    xlsx_path = Path(config.folders.results_folder) / f"extracted_{file_path.stem}.xlsx"
                    self._save_records_to_excel(extraction_result.records, xlsx_path)
                    logger.info(f"  Excel saved: {xlsx_path.name}")
                except ImportError:
                    logger.debug("openpyxl not installed, skipping Excel export")
                except Exception as e:
                    logger.warning(f"Failed to save Excel: {e}")

                # Move processed file to results folder
                self._move_processed_file(file_path)

            else:
                logger.error(f"Processing failed: No records extracted")

        except Exception as e:
            logger.error(f"Error processing {file_path.name}: {e}", exc_info=True)

    def _move_processed_file(self, file_path: Path):
        """Move processed PDF file to results folder with 'Processed_' prefix"""
        try:
            results_folder = Path(config.folders.results_folder)
            new_filename = f"Processed_{file_path.name}"
            destination = results_folder / new_filename

            # Handle duplicate filenames
            counter = 1
            while destination.exists():
                new_filename = f"Processed_{file_path.stem}_{counter}{file_path.suffix}"
                destination = results_folder / new_filename
                counter += 1

            import shutil
            shutil.move(str(file_path), str(destination))
            logger.info(f"  Moved to: {destination}")

        except Exception as e:
            logger.warning(f"Failed to move processed file: {e}")

    def _process_with_mistral(self, file_path: Path):
        """Process using Mistral for both OCR and extraction"""
        from mistral_processor import MistralProcessor, DocumentCategory

        category_map = {
            1: DocumentCategory.EOB,
            2: DocumentCategory.FACESHEET,
            3: DocumentCategory.INVOICE
        }
        category = category_map.get(self.doc_category, DocumentCategory.EOB)

        # Use batch processing setting from config
        use_batch = config.processing.use_batch_processing
        processor = MistralProcessor(use_batch=use_batch)
        result = processor.process_pdf(str(file_path), category, file_path.name)

        # Get page count
        from PyPDF2 import PdfReader
        reader = PdfReader(str(file_path))
        total_pages = len(reader.pages)

        return result, total_pages

    def _process_mistral_ocr_openai_llm(self, file_path: Path):
        """Process using Mistral OCR + OpenAI for extraction (with chunking for large docs)"""
        from mistral_processor import MistralProcessor
        from openai_extractor import OpenAIExtractor, DocumentCategory
        from PyPDF2 import PdfReader, PdfWriter
        import tempfile
        import os

        category_map = {
            1: DocumentCategory.EOB,
            2: DocumentCategory.FACESHEET,
            3: DocumentCategory.INVOICE
        }
        category = category_map.get(self.doc_category, DocumentCategory.EOB)

        # Get page count
        reader = PdfReader(str(file_path))
        total_pages = len(reader.pages)

        logger.info(f"Hybrid mode: Mistral OCR + OpenAI LLM")
        logger.info(f"Document has {total_pages} pages")

        MAX_PAGES_PER_CHUNK = 30
        mistral = MistralProcessor()
        openai_extractor = OpenAIExtractor()

        if total_pages <= MAX_PAGES_PER_CHUNK:
            # Small document - process directly
            logger.info("Processing document in single pass...")
            raw_text = mistral.get_raw_text(str(file_path))
            raw_data = {
                'text': raw_text,
                'pages': total_pages,
                'page_details': []
            }
            result = openai_extractor.extract_data(raw_data, category, file_path.name)
            return result, total_pages

        # Large document - process in chunks
        num_chunks = (total_pages + MAX_PAGES_PER_CHUNK - 1) // MAX_PAGES_PER_CHUNK
        logger.info(f"Large document: splitting into {num_chunks} chunks")

        all_records = []
        total_input_tokens = 0
        total_output_tokens = 0

        for chunk_idx in range(num_chunks):
            start_page = chunk_idx * MAX_PAGES_PER_CHUNK
            end_page = min(start_page + MAX_PAGES_PER_CHUNK, total_pages)

            logger.info(f"Processing chunk {chunk_idx + 1}/{num_chunks} (pages {start_page + 1}-{end_page})")

            # Create temporary PDF for this chunk
            writer = PdfWriter()
            for page_num in range(start_page, end_page):
                writer.add_page(reader.pages[page_num])

            with tempfile.NamedTemporaryFile(suffix='.pdf', delete=False) as tmp_file:
                writer.write(tmp_file)
                tmp_path = tmp_file.name

            try:
                # OCR with Mistral
                raw_text = mistral.get_raw_text(tmp_path)

                # Wrap text in expected format for OpenAI extractor
                raw_data = {
                    'text': raw_text,
                    'pages': end_page - start_page,
                    'page_details': []
                }

                # Extract with OpenAI
                chunk_result = openai_extractor.extract_data(
                    raw_data, category, f"{file_path.name}_chunk{chunk_idx + 1}"
                )

                # Adjust page numbers
                for record in chunk_result.records:
                    if 'Original_page_no' in record:
                        original = record.get('Original_page_no', 1)
                        if isinstance(original, int):
                            record['Original_page_no'] = start_page + original
                    else:
                        record['Original_page_no'] = start_page + 1

                all_records.extend(chunk_result.records)
                total_input_tokens += chunk_result.input_tokens
                total_output_tokens += chunk_result.output_tokens

                logger.info(f"Chunk {chunk_idx + 1}: extracted {len(chunk_result.records)} records")

            finally:
                os.unlink(tmp_path)

        # Create combined result
        from openai_extractor import ExtractionResult
        combined_result = ExtractionResult(
            data={'chunks': num_chunks, 'total_records': len(all_records)},
            records=all_records,
            input_tokens=total_input_tokens,
            output_tokens=total_output_tokens,
            total_tokens=total_input_tokens + total_output_tokens,
            model_used=config.openai.model,
            validation_passed=True,
            confidence_score=0.9
        )

        logger.info(f"Total records extracted: {len(all_records)}")
        return combined_result, total_pages

    def _process_with_google_docai(self, file_path: Path, use_mistral_llm: bool = False):
        """Process using Google Document AI for OCR"""
        from document_ai_processor import EOBProcessor

        category_map = {
            1: 'EOB',
            2: 'FACESHEET',
            3: 'INVOICE'
        }

        eob_processor = EOBProcessor()
        docai_result = eob_processor.process_file(str(file_path))

        if not docai_result or docai_result.get('status') != 'success':
            raise Exception(f"Document AI processing failed: {docai_result.get('error', 'Unknown error')}")

        raw_data = docai_result['raw_data']
        total_pages = docai_result['total_pages']

        if use_mistral_llm:
            # Use Mistral for extraction
            from mistral_processor import MistralProcessor, DocumentCategory
            category_map_enum = {
                1: DocumentCategory.EOB,
                2: DocumentCategory.FACESHEET,
                3: DocumentCategory.INVOICE
            }
            category = category_map_enum.get(self.doc_category, DocumentCategory.EOB)

            # Create a simple extraction result from text
            # Note: This is a hybrid mode - might be less efficient
            processor = MistralProcessor()
            # For text-based extraction, we'd need to modify the processor
            # For now, fall back to OpenAI in this edge case
            logger.warning("Hybrid Google DocAI + Mistral LLM not fully supported, using OpenAI")
            from openai_extractor import OpenAIExtractor, DocumentCategory as OAICategory
            category = {1: OAICategory.EOB, 2: OAICategory.FACESHEET, 3: OAICategory.INVOICE}.get(
                self.doc_category, OAICategory.EOB
            )
            extractor = OpenAIExtractor()
            result = extractor.extract_data(raw_data, category, file_path.name)
        else:
            # Use OpenAI for extraction
            from openai_extractor import OpenAIExtractor, DocumentCategory
            category = {1: DocumentCategory.EOB, 2: DocumentCategory.FACESHEET, 3: DocumentCategory.INVOICE}.get(
                self.doc_category, DocumentCategory.EOB
            )
            extractor = OpenAIExtractor()
            result = extractor.extract_data(raw_data, category, file_path.name)

        return result, total_pages

    # Define standard column order for EOB documents
    EOB_COLUMN_ORDER = [
        "Original_page_no",
        "EOB_page_no",
        "patient_acct",
        "Patient_ID",
        "Claim_ID",
        "Patient Name",
        "First_Name",
        "Last Name",
        "member_number",
        "service_date",
        "allowed_amount",
        "interest_amount",
        "paid_amount",
        "insurance_co",
        "billed_amount",
        "cpt_hcpcs",
        "adj_co45",
        "adj_co144",
        "adj_co253",
        "check_number",
        "account_number",
        "patient_responsibility",
        "claim_summary",
        "action_required",
        "reason_code_comments",
        "Confidence_Score"
    ]

    def _get_ordered_headers(self, records: list) -> list:
        """Get headers in the correct order for EOB documents"""
        # Get all unique keys from all records
        all_keys = set()
        for record in records:
            if isinstance(record, dict):
                all_keys.update(record.keys())

        # Start with EOB column order, then add any extra columns at the end
        ordered_headers = []
        for col in self.EOB_COLUMN_ORDER:
            if col in all_keys:
                ordered_headers.append(col)
                all_keys.discard(col)

        # Add any remaining columns (not in standard order) at the end, sorted
        ordered_headers.extend(sorted(all_keys))

        return ordered_headers

    def _save_records_to_csv(self, records: list, output_path: Path):
        """Save records to CSV file"""
        import csv

        if not records:
            return

        headers = self._get_ordered_headers(records)

        with open(output_path, 'w', newline='', encoding='utf-8') as f:
            writer = csv.DictWriter(f, fieldnames=headers, extrasaction='ignore')
            writer.writeheader()

            for record in records:
                if isinstance(record, dict):
                    # Clean values for CSV
                    clean_record = {}
                    for key, value in record.items():
                        if isinstance(value, (list, dict)):
                            clean_record[key] = json.dumps(value)
                        else:
                            clean_record[key] = value
                    writer.writerow(clean_record)

    def _save_records_to_excel(self, records: list, output_path: Path):
        """Save records to Excel file"""
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

        if not records:
            return

        # Use ordered headers
        headers = self._get_ordered_headers(records)

        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Extracted Data"

        # Style definitions
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Write headers
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
            cell.border = thin_border

        # Write data rows
        for row_num, record in enumerate(records, 2):
            if isinstance(record, dict):
                for col, header in enumerate(headers, 1):
                    value = record.get(header, '')
                    if isinstance(value, (list, dict)):
                        value = json.dumps(value)
                    cell = ws.cell(row=row_num, column=col, value=value)
                    cell.border = thin_border

        # Auto-adjust column widths
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)  # Cap at 50
            ws.column_dimensions[column].width = adjusted_width

        # Freeze header row
        ws.freeze_panes = 'A2'

        wb.save(output_path)


class FolderWatcher:
    """Watch a folder for new PDF files"""

    def __init__(self, watch_folder: str = None, doc_category: int = 1):
        self.watch_folder = watch_folder or config.folders.upload_folder
        self.doc_category = doc_category
        self.observer = None

    def start(self):
        """Start watching the folder"""
        if not Path(self.watch_folder).exists():
            logger.error(f"Watch folder does not exist: {self.watch_folder}")
            return

        logger.info(f"Starting folder watcher...")
        logger.info(f"  Watch Folder: {self.watch_folder}")
        logger.info(f"  Results Folder: {config.folders.results_folder}")
        logger.info(f"  Document Category: {self.doc_category}")

        event_handler = PDFHandler(self.doc_category)

        # Process existing files first
        self._process_existing_files(event_handler)

        # Then start watching for new files
        self.observer = Observer()
        self.observer.schedule(event_handler, self.watch_folder, recursive=False)
        self.observer.start()

        logger.info("Folder watcher started. Press Ctrl+C to stop.")

        try:
            while True:
                time.sleep(1)
        except KeyboardInterrupt:
            self.stop()

    def _process_existing_files(self, handler: PDFHandler):
        """Process any existing PDF files in the watch folder"""
        watch_path = Path(self.watch_folder)
        existing_pdfs = list(watch_path.glob("*.pdf"))

        # Filter out already processed files
        unprocessed = [
            f for f in existing_pdfs
            if not f.name.startswith("Processed_")
        ]

        if unprocessed:
            logger.info(f"Found {len(unprocessed)} existing PDF file(s) to process")
            for pdf_file in unprocessed:
                logger.info(f"Processing existing file: {pdf_file.name}")
                handler._process_file(pdf_file)
        else:
            logger.info("No existing unprocessed PDF files found")

    def stop(self):
        """Stop watching the folder"""
        if self.observer:
            logger.info("Stopping folder watcher...")
            self.observer.stop()
            self.observer.join()
            logger.info("Folder watcher stopped.")


def watch_folder(folder: str = None, category: int = 1):
    """Convenience function to start folder watching"""
    watcher = FolderWatcher(folder, category)
    watcher.start()


if __name__ == "__main__":
    import argparse

    parser = argparse.ArgumentParser(description='Watch folder for new PDF files')
    parser.add_argument('--folder', '-f', help='Folder to watch (default: upload folder from config)')
    parser.add_argument('--category', '-c', type=int, default=1,
                        choices=[1, 2, 3],
                        help='Document category (1=EOB, 2=Facesheet, 3=Invoice)')

    args = parser.parse_args()

    logging.basicConfig(
        level=logging.INFO,
        format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
    )

    watch_folder(args.folder, args.category)
